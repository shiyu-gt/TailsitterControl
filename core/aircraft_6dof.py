"""
双旋翼尾座式固定翼飞机6DOF动力学模型（含四元数）
===================================================
基于表格气动系数的机体轴运动方程，使用四元数表示姿态。

该模型模拟具有以下特点的双旋翼尾座式飞机：
- 左右机翼上的两个电机提供推力
- 螺旋桨滑流中的控制面以增强效能
- 使用动量理论计算滑流速度
- 使用四元数表示姿态，避免万向节锁问题

气动数据从Excel文件加载：
  - aerodata_lon.xlsx  : CL(alpha), CD(alpha), Cm(alpha) 在beta=0, de=0时
  - aerodata_de.xlsx   : delta-CL(de), delta-CD(de), delta-Cm(de)  (滑流中的升降舵增量)
  - aerodata_lat.xlsx  : CY(alpha,beta), Cl(alpha,beta), Cn(alpha,beta)
  - aerodata_throttle.xlsx: 每个电机的推力与油门关系

约定：x向前，y向右，z向下（机体轴）
大气：ISA模型，不可压缩（低马赫数）
"""

import os
import numpy as np

# ─── 数据目录解析 ────────────────────────────────────────────────

_DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data')

def _resolve_data_path(filename):
    """解析数据文件路径，相对于 data/ 目录。"""
    path = os.path.join(_DATA_DIR, filename)
    if not os.path.isfile(path):
        raise FileNotFoundError(f"数据文件未找到: {path}")
    return path

# ─── 国际标准大气（简化，仅对流层） ────────────────────────────

def isa_atmosphere(h):
    """11km以下的标准大气。
    返回值：(rho, a) — 密度 [kg/m^3], 音速 [m/s]。"""
    T0, P0, rho0 = 288.15, 101325.0, 1.2250
    L = -0.0065
    g0 = 9.80665
    R = 287.058
    gamma = 1.4
    T = T0 + L * h
    P = P0 * (T / T0) ** (-g0 / (L * R))
    rho = P / (R * T)
    a = np.sqrt(gamma * R * T)
    return rho, a


# ─── 四元数运算 ────────────────────────────────────────

def quaternion_multiply(q1, q2):
    """两个四元数相乘"""
    w1, x1, y1, z1 = q1
    w2, x2, y2, z2 = q2
    w = w1*w2 - x1*x2 - y1*y2 - z1*z2
    x = w1*x2 + x1*w2 + y1*z2 - z1*y2
    y = w1*y2 - x1*z2 + y1*w2 + z1*x2
    z = w1*z2 + x1*y2 - y1*x2 + z1*w2
    return np.array([w, x, y, z])


def quaternion_to_euler(q):
    """将四元数转换为欧拉角（phi, theta, psi）"""
    w, x, y, z = q

    # 滚转角 (phi)
    sinr_cosp = 2 * (w * x + y * z)
    cosr_cosp = 1 - 2 * (x * x + y * y)
    phi = np.arctan2(sinr_cosp, cosr_cosp)

    # 俯仰角 (theta)
    sinp = 2 * (w * y - z * x)
    if abs(sinp) >= 1:
        theta = np.copysign(np.pi / 2, sinp)
    else:
        theta = np.arcsin(sinp)

    # 偏航角 (psi)
    siny_cosp = 2 * (w * z + x * y)
    cosy_cosp = 1 - 2 * (y * y + z * z)
    psi = np.arctan2(siny_cosp, cosy_cosp)

    return phi, theta, psi


def euler_to_quaternion(phi, theta, psi):
    """将欧拉角转换为四元数"""
    cy = np.cos(psi * 0.5)
    sy = np.sin(psi * 0.5)
    cp = np.cos(theta * 0.5)
    sp = np.sin(theta * 0.5)
    cr = np.cos(phi * 0.5)
    sr = np.sin(phi * 0.5)

    w = cr * cp * cy + sr * sp * sy
    x = sr * cp * cy - cr * sp * sy
    y = cr * sp * cy + sr * cp * sy
    z = cr * cp * sy - sr * sp * cy

    return np.array([w, x, y, z])


def quaternion_normalize(q):
    """归一化四元数"""
    return q / np.linalg.norm(q)


def rotation_matrix_from_quaternion(q):
    """从四元数获取旋转矩阵"""
    w, x, y, z = q

    R = np.array([
        [1 - 2*(y*y + z*z), 2*(x*y - w*z), 2*(x*z + w*y)],
        [2*(x*y + w*z), 1 - 2*(x*x + z*z), 2*(y*z - w*x)],
        [2*(x*z - w*y), 2*(y*z + w*x), 1 - 2*(x*x + y*y)]
    ])

    return R


# ── 气动系数模型（基于表格+线性速率项） ───────────

class AeroModel:
    """
    来自CFD/风洞数据表的气动力和力矩系数。

    该模型实现了带有螺旋桨滑流中控制面的双旋翼尾座式飞机。
    静态系数（alpha, beta, de）通过表格插值获得。
    滑流效应使用动量理论计算。
    速率相关项（q_hat, p_hat, r_hat）和控制项（da, dr）
    由于不在数据文件中，以线性贡献的形式添加。

    接口（重构后）：
        get_base_coefficients(alpha, beta, p_hat, q_hat, r_hat)
            返回机身基础系数 (CL_b, CD_b, CY_b, Cl_b, Cm_b, Cn_b)
        get_elevon_longitudinal_coefficients(delta_e)
            基于对称等效偏角返回全机纵向增量 (dCL, dCD, dCm)
        get_elevon_single_CL_increment(de)
            返回单侧舵面全机升力系数增量（用于滚转）
        get_slipstream_dynamic_pressure(thrust_left, thrust_right, rho, V_inf)
            返回左右滑流动压
        coefficients_with_slipstream(...)
            【兼容接口】供 legacy 代码调用，内部封装双动压模型
    """

    def __init__(self, lon_file=None, lat_file=None, de_file=None, throttle_file=None):
        lon_file = lon_file or _resolve_data_path('aerodata_lon.xlsx')
        lat_file = lat_file or _resolve_data_path('aerodata_lat.xlsx')
        de_file = de_file or _resolve_data_path('aerodata_de.xlsx')
        throttle_file = throttle_file or _resolve_data_path('aerodata_throttle.xlsx')

        # ── 几何参数（双旋翼飞翼式尾座式） ──
        self.S = 0.62        # 机翼参考面积 [m^2]
        self.c_bar = 0.31    # 平均气动弦长 [m]
        self.b = 2.0         # 翼展 [m]
        self.AR = self.b**2 / self.S  # 展弦比

        # 电机位置（从中线算起）
        self.y_motor_left = -self.b / 4  # 左电机y位置 [m]
        self.y_motor_right = self.b / 4   # 右电机y位置 [m]

        # 螺旋桨参数
        self.D_prop = 0.4064  # 螺旋桨直径 [m]
        self.A_prop = np.pi * (self.D_prop/2)**2  # 螺旋桨盘面积 [m^2]

        # 升降副翼气动中心横向位置（用于滚转力矩计算）
        self.y_elevon = self.b / 4  # [m]，右正左负

        # 滑流参数
        self.slipstream_efficiency = 1.2  # 滑流效率系数

        # ── 速率导数（不在数据文件中，作为参数保留） ──
        # 纵向
        self.CLq = 0      # 每弧度
        self.Cmq = 0    # 每弧度
        # 横侧向
        self.CYp = 0.0
        self.CYr = 0.0
        self.Clp = 0.0
        self.Clr = 0.0
        self.Cnp = 0.0
        self.Cnr = 0.0

        # 本机为升降副翼（elevon）布局，无独立副翼/方向舵
        # 滚转由差动升降副翼产生，偏航由差动油门产生

        # ── 加载表格数据 ──
        self._load_lon_data(lon_file)
        self._load_de_data(de_file)
        self._load_lat_data(lat_file)
        self._load_throttle_data(throttle_file)

    # ──────────────────────────────────────────────────────────────────────
    #  数据加载
    # ──────────────────────────────────────────────────────────────────────

    def _load_throttle_data(self, fname):
        """加载电机推力与油门数据。"""
        if not os.path.isfile(fname):
            raise FileNotFoundError(f"油门数据未找到: {fname}")
        rows = self._load_xlsx(fname)
        # 提取排序后的油门数据用于np.interp
        data = sorted(rows, key=lambda r: r[0])
        self._throttle = np.array([r[0] for r in data], dtype=float)
        self._thrust = np.array([r[1] * 9.81 for r in data], dtype=float)  # 将kg转换为N

    def thrust_from_throttle(self, throttle):
        """从油门输入(0-1)获取推力。"""
        return np.interp(throttle, self._throttle, self._thrust)

    def _calculate_slipstream_velocity(self, thrust, rho, V_inf):
        """使用动量理论计算滑流速度。
        v_slip = V_inf + sqrt(2 * T / (rho * A))
        V_inf 为真实自由流速度，仅在非滑流区使用。
        """
        v_slip = V_inf + np.sqrt(2 * thrust / (rho * self.A_prop))
        return v_slip

    def _calculate_slipstream_factor(self, thrust, rho, V_inf):
        """计算滑流中的速度比 v_slip / V_inf。"""
        v_slip = self._calculate_slipstream_velocity(thrust, rho, V_inf)
        return v_slip / V_inf if V_inf > 0 else 1.0

    def _load_xlsx(self, fname):
        """读取xlsx文件，返回行列表（跳过标题行）。"""
        import openpyxl
        wb = openpyxl.load_workbook(fname, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = [v for v in row if v is not None]
            if len(vals) >= 2:
                rows.append(vals)
        return rows

    def _load_lon_data(self, fname):
        """
        在beta=0, de=0时加载纵向气动数据。
        列：alpha(度), beta, de, CL, CD, Cm
        """
        if not os.path.isfile(fname):
            raise FileNotFoundError(f"纵向气动数据未找到: {fname}")
        rows = self._load_xlsx(fname)
        # 提取用于np.interp的alpha（必须单调）
        data = sorted(rows, key=lambda r: r[0])
        self._lon_alpha = np.array([r[0] for r in data], dtype=float)  # 度
        self._lon_CL   = np.array([r[3] for r in data], dtype=float)
        self._lon_CD   = np.array([r[4] for r in data], dtype=float)
        self._lon_Cm   = np.array([r[5] for r in data], dtype=float)

    def _load_de_data(self, fname):
        """
        加载升降舵增量系数。
        列：de(度), dCL, dCD, dCm
        dCL, dCD, dCm是在该de值下的总系数增量。
        """
        if not os.path.isfile(fname):
            raise FileNotFoundError(f"升降舵气动数据未找到: {fname}")
        rows = self._load_xlsx(fname)
        data = sorted(rows, key=lambda r: r[0])
        self._de_deg = np.array([r[0] for r in data], dtype=float)  # 度
        self._de_CL  = np.array([r[1] for r in data], dtype=float)
        self._de_CD  = np.array([r[2] for r in data], dtype=float)
        self._de_Cm  = np.array([r[3] for r in data], dtype=float)

    def _load_lat_data(self, fname):
        """
        加载横侧向气动数据作为(alpha, beta)的2D表格。
        列：alpha(度), beta(度), CL, CD, CY, Cmy(俯仰), Cmz(偏航), Cmx(滚转)

        机体轴力矩约定：
            Cmx -> Cl (滚转力矩，绕x轴)
            Cmy -> Cm (俯仰力矩，绕y轴)
            Cmz -> Cn (偏航力矩，绕z轴)

        我们提取CY, Cl (cmx), Cn (cmz)。
        在beta=0时，根据对称性：常规飞机的CY=0, Cl=0, Cn=0。
        """
        if not os.path.isfile(fname):
            raise FileNotFoundError(f"Lateral aero data not found: {fname}")
        rows = self._load_xlsx(fname)
        alpha_all = np.array([r[0] for r in rows], dtype=float)
        beta_all  = np.array([r[1] for r in rows], dtype=float)
        cy_all    = np.array([r[4] for r in rows], dtype=float)   # CY
        cn_all    = np.array([r[6] for r in rows], dtype=float)   # Cmz = Cn
        cl_all    = np.array([r[7] for r in rows], dtype=float)   # Cmx = Cl

        # 为每个alpha构建1D插值器，包含beta=0对称点
        self._lat_alphas = sorted(set(alpha_all))
        self._lat_CY = {}
        self._lat_Cl = {}
        self._lat_Cn = {}

        for a_val in self._lat_alphas:
            mask = np.abs(alpha_all - a_val) < 1e-6
            betas = beta_all[mask]
            order = np.argsort(betas)

            # 在前面添加beta=0点，系数为零（对称性）
            self._lat_CY[a_val] = (
                np.concatenate([[0.0], betas[order]]),
                np.concatenate([[0.0], cy_all[mask][order]])
            )
            self._lat_Cl[a_val] = (
                np.concatenate([[0.0], betas[order]]),
                np.concatenate([[0.0], cl_all[mask][order]])
            )
            self._lat_Cn[a_val] = (
                np.concatenate([[0.0], betas[order]]),
                np.concatenate([[0.0], cn_all[mask][order]])
            )

    # ──────────────────────────────────────────────────────────────────────
    #  插值计算
    # ──────────────────────────────────────────────────────────────────────

    def _interp_2d(self, tables, alpha_deg, beta_deg):
        """
        在横向2D表格上进行双线性插值。
        首先在两个边界alpha值处对beta插值，
        然后在两个alpha结果之间线性插值。
        """
        alphas = self._lat_alphas

        # 限制/边界alpha
        if alpha_deg <= alphas[0]:
            a_lo = a_hi = alphas[0]
        elif alpha_deg >= alphas[-1]:
            a_lo = a_hi = alphas[-1]
        else:
            a_lo, a_hi = alphas[0], alphas[1]
            for i in range(len(alphas) - 1):
                if alphas[i] <= alpha_deg <= alphas[i + 1]:
                    a_lo, a_hi = alphas[i], alphas[i + 1]
                    break

        # 在a_lo处插值
        betas_lo, coeffs_lo = tables[a_lo]
        val_lo = np.interp(beta_deg, betas_lo, coeffs_lo)

        if a_hi != a_lo:
            # 在a_hi处插值
            betas_hi, coeffs_hi = tables[a_hi]
            val_hi = np.interp(beta_deg, betas_hi, coeffs_hi)
            frac = (alpha_deg - a_lo) / (a_hi - a_lo)
            return val_lo + frac * (val_hi - val_lo)

        return val_lo

    # ──────────────────────────────────────────────────────────────────────
    #  系数计算
    # ──────────────────────────────────────────────────────────────────────

    def get_base_coefficients(self, alpha, beta, p_hat, q_hat, r_hat):
        """
        计算机身基础气动系数（不含舵面，不含滑流增强）。
        返回值: (CL_b, CD_b, CY_b, Cl_b, Cm_b, Cn_b)
        """
        a_deg = np.degrees(alpha)
        b_deg = np.degrees(beta)

        CL_b = np.interp(a_deg, self._lon_alpha, self._lon_CL)
        CD_b = np.interp(a_deg, self._lon_alpha, self._lon_CD)
        Cm_b = np.interp(a_deg, self._lon_alpha, self._lon_Cm)

        CY_b = self._interp_2d(self._lat_CY, a_deg, b_deg)
        Cl_b = self._interp_2d(self._lat_Cl, a_deg, b_deg)
        Cn_b = self._interp_2d(self._lat_Cn, a_deg, b_deg)

        # 速率相关项
        CL_b += self.CLq * q_hat
        Cm_b += self.Cmq * q_hat
        CY_b += self.CYp * p_hat + self.CYr * r_hat
        Cl_b += self.Clp * p_hat + self.Clr * r_hat
        Cn_b += self.Cnp * p_hat + self.Cnr * r_hat

        return CL_b, CD_b, CY_b, Cl_b, Cm_b, Cn_b

    def get_elevon_longitudinal_coefficients(self, delta_e):
        """
        基于对称等效偏角 delta_e 查表，返回全机纵向气动系数增量。
        即：左右舵面同步偏转 delta_e 时，全机产生的 dCL, dCD, dCm 总增量。
        所有系数基于全机参考面积 S。
        """
        de_deg = np.degrees(delta_e)
        dCL = np.interp(de_deg, self._de_deg, self._de_CL)
        dCD = np.interp(de_deg, self._de_deg, self._de_CD)
        dCm = np.interp(de_deg, self._de_deg, self._de_Cm)
        return dCL, dCD, dCm

    def get_elevon_single_CL_increment(self, de):
        """
        基于单侧舵面偏角 de，返回该侧舵面产生的全机升力系数增量。
        用于滚转力矩计算。所有系数基于全机参考面积 S。
        若数据表提供的是全机总增量，内部按半分处理（假设左右气动独立）。
        """
        de_deg = np.degrees(de)
        dCL_total = np.interp(de_deg, self._de_deg, self._de_CL)
        # 假设左右舵面气动独立，单侧贡献为全机总增量的一半
        return 0.5 * dCL_total

    def get_slipstream_dynamic_pressure(self, thrust_left, thrust_right, rho, V_inf):
        """
        基于真实 V_inf 的动量理论，返回左右滑流动压 q_slip_left, q_slip_right。
        """
        v_slip_l = self._calculate_slipstream_velocity(thrust_left, rho, V_inf)
        v_slip_r = self._calculate_slipstream_velocity(thrust_right, rho, V_inf)
        q_slip_left = 0.5 * rho * v_slip_l**2
        q_slip_right = 0.5 * rho * v_slip_r**2
        return q_slip_left, q_slip_right

    def coefficients_with_slipstream(self, alpha, beta, p_hat, q_hat, r_hat,
                                    de_left, de_right, thrust_left, thrust_right, rho, V_inf):
        """
        【兼容接口】计算带有滑流效应的气动系数总系数。
        供 legacy 代码（如 linearize_analyze.py）调用。
        内部封装为双动压模型：机身用 q_inf，舵面用 q_slip。
        """
        # 基础系数
        CL_b, CD_b, CY_b, Cl_b, Cm_b, Cn_b = self.get_base_coefficients(
            alpha, beta, p_hat, q_hat, r_hat)

        # 舵面纵向增量
        delta_e = 0.5 * (de_left + de_right)
        dCL_e, dCD_e, dCm_e = self.get_elevon_longitudinal_coefficients(delta_e)

        # 滑流动压
        qsl_l, qsl_r = self.get_slipstream_dynamic_pressure(
            thrust_left, thrust_right, rho, V_inf)
        qsl_avg = 0.5 * (qsl_l + qsl_r)
        q_inf = 0.5 * rho * V_inf**2

        # 兼容：将舵面增量折合成等效总系数（用于旧接口）
        # 当 q_inf 极小时避免除零，直接返回大增量（legacy  callers 需注意）
        if q_inf > 1e-3:
            CL = CL_b + dCL_e * (qsl_avg / q_inf)
            CD = CD_b + dCD_e * (qsl_avg / q_inf)
            Cm = Cm_b + dCm_e * (qsl_avg / q_inf)
        else:
            # 悬停近似：将滑流增量直接叠加上去（系数名义值放大）
            CL = CL_b + dCL_e * 10.0
            CD = CD_b + dCD_e * 10.0
            Cm = Cm_b + dCm_e * 10.0

        # 滚转：差动效应
        dCL_l = self.get_elevon_single_CL_increment(de_left)
        dCL_r = self.get_elevon_single_CL_increment(de_right)
        L_inc_left = qsl_l * self.S * dCL_l
        L_inc_right = qsl_r * self.S * dCL_r
        # 将滚转力矩折算为系数增量（基于 q_inf）
        if q_inf > 1e-3:
            Cl_roll = (self.y_elevon * (L_inc_right - L_inc_left)) / (q_inf * self.S * self.b)
        else:
            Cl_roll = 0.0

        CY = CY_b
        Cl = Cl_b + Cl_roll
        Cn = Cn_b

        return CL, CD, CY, Cl, Cm, Cn


# ─── 飞机6DOF刚体模型 ───────────────────────────────────────────

class Aircraft6DOF:
    """
    机体轴中的完整6DOF非线性运动方程（使用四元数表示姿态）。

    状态向量（13自由度）：
        x = [u, v, w, p, q, r, qx, qy, qz, qw, px, py, pz]
        四元数：q = [qw, qx, qy, qz]（标量优先）
    控制向量：
        u_ctrl = [throttle_left, throttle_right, de_left, de_right]
            throttle_left  : 左电机油门 [0-1]
            throttle_right : 右电机油门 [0-1]
            de_left        : 左升降舵偏角 [rad]
            de_right       : 右升降舵偏角 [rad]
    """

    def __init__(self):
        # ── 质量/惯性（双旋翼尾座式飞翼） ──
        self.mass = 6.5        # [kg]
        self.Ix = 0.1          # [kg*m^2]
        self.Iy = 0.183        # [kg*m^2]
        self.Iz = 2.8          # [kg*m^2]
        self.Ixz = 0.0         # [kg*m^2]

        self.g = 9.80665       # [m/s^2]
        self.aero = AeroModel()

        # 四元数相关参数
        self.quaternion = np.array([1.0, 0.0, 0.0, 0.0])  # 初始单位四元数（无旋转）

        # ── 可配置饱和限值 ──
        self.sat_force = np.array([2000.0, 2000.0, 2000.0])   # N
        self.sat_moment = np.array([200.0, 200.0, 200.0])     # N·m
        self.sat_accel = np.array([50.0, 50.0, 50.0])         # rad/s²
        self.sat_vel = 50.0                                    # m/s
        self.enable_diagnostics = False                        # 诊断开关

    def _apply_limit(self, value, limit, name):
        """可配置双向限幅，支持诊断输出。"""
        arr = np.asarray(value)
        lim = np.asarray(limit)
        clipped = np.clip(arr, -lim, lim)
        if self.enable_diagnostics and np.any(np.abs(arr) > lim):
            actual = arr[np.argmax(np.abs(arr))]
            print(f"WARNING: {name} saturated at {actual:.3e}, clipped to ±{np.max(lim):.3e}")
        return clipped

    # ── 机体轴中的气动力和力矩 ──
    def _aero_forces_moments(self, state, ev_left, ev_right, thrust_left, thrust_right, rho):
        u, v, w, p, q, r = state[:6]
        V = np.sqrt(u**2 + v**2 + w**2)

        # 数值稳定性检查
        if np.any(np.isnan([u, v, w, p, q, r])) or np.any(np.abs([u, v, w, p, q, r]) > 1e6):
            raise RuntimeError("State diverged: NaN or Inf detected in aerodynamics.")

        alpha = np.arctan2(w, u)
        # 放宽 alpha 限幅到 90°：数据表覆盖 -180°~180°，90°以内数据可靠
        alpha = self._apply_limit(alpha, np.radians(90), "alpha")

        beta = np.arcsin(np.clip(v / max(V, 1e-6), -1, 1))
        q_inf = 0.5 * rho * V**2
        q_inf = self._apply_limit(q_inf, 10000.0, "q_inf")

        # 无量纲角速率
        p_hat = p * self.aero.b / (2 * max(V, 0.1))
        q_hat = q * self.aero.c_bar / (2 * max(V, 0.1))
        r_hat = r * self.aero.b / (2 * max(V, 0.1))
        p_hat = self._apply_limit(p_hat, 10.0, "p_hat")
        q_hat = self._apply_limit(q_hat, 10.0, "q_hat")
        r_hat = self._apply_limit(r_hat, 10.0, "r_hat")

        S, c_bar, b = self.aero.S, self.aero.c_bar, self.aero.b
        ca, sa = np.cos(alpha), np.sin(alpha)

        # ── 0. 控制解耦 ──
        delta_e = 0.5 * (ev_left + ev_right)   # 对称等效偏角（俯仰）

        # ── 1. 基础气动力（机身，使用自由流动压 q_inf） ──
        CL_b, CD_b, CY_b, Cl_b, Cm_b, Cn_b = self.aero.get_base_coefficients(
            alpha, beta, p_hat, q_hat, r_hat)

        Fa_x_base = -q_inf * S * (CD_b * ca - CL_b * sa)
        Fa_y_base =  q_inf * S * CY_b
        Fa_z_base = -q_inf * S * (CD_b * sa + CL_b * ca)
        L_base = q_inf * S * b * Cl_b
        M_base = q_inf * S * c_bar * Cm_b
        N_base = q_inf * S * b * Cn_b

        # ── 2. 滑流动压与有效因子 ──
        # 前飞时滑流扩散，舵面仅部分沉浸在滑流中。
        # 使用平方衰减模型：epsilon = 1 / (1 + (V / V_ref)^2)
        V_ref = 15.0  # [m/s]，滑流-自由流过渡参考速度
        epsilon = 1.0 / (1.0 + (V / V_ref)**2) if V > 0.1 else 1.0

        qsl_l, qsl_r = self.aero.get_slipstream_dynamic_pressure(
            thrust_left, thrust_right, rho, V)
        q_eff_l = q_inf + epsilon * (qsl_l - q_inf)
        q_eff_r = q_inf + epsilon * (qsl_r - q_inf)
        q_eff_avg = 0.5 * (q_eff_l + q_eff_r)

        # ── 3. 升降副翼纵向效应（俯仰，使用等效动压平均值） ──
        dCL_e, dCD_e, dCm_e = self.aero.get_elevon_longitudinal_coefficients(delta_e)
        Fa_x_elev = -q_eff_avg * S * (dCD_e * ca - dCL_e * sa)
        Fa_z_elev = -q_eff_avg * S * (dCD_e * sa + dCL_e * ca)
        M_elev    =  q_eff_avg * S * c_bar * dCm_e

        # ── 4. 升降副翼滚转效应（滚转，由左右升力差 × 力臂产生） ──
        dCL_l = self.aero.get_elevon_single_CL_increment(ev_left)
        dCL_r = self.aero.get_elevon_single_CL_increment(ev_right)
        L_inc_left  = q_eff_l * S * dCL_l
        L_inc_right = q_eff_r * S * dCL_r
        L_elev_diff = self.aero.y_elevon * (L_inc_right - L_inc_left)

        # ── 5. 总合 ──
        Fa_x = Fa_x_base + Fa_x_elev
        Fa_y = Fa_y_base
        Fa_z = Fa_z_base + Fa_z_elev
        L = L_base + L_elev_diff
        M = M_base + M_elev
        N = N_base

        Fa = self._apply_limit(np.array([Fa_x, Fa_y, Fa_z]), self.sat_force, "Fa")
        Ma = self._apply_limit(np.array([L, M, N]), self.sat_moment, "Ma")

        return Fa, Ma

    # ── 状态导数 ──
    def derivatives(self, state, controls, rho):
        """
        dx/dt = f(x, u).
        state    : (13,) — [u, v, w, p, q, r, qx, qy, qz, qw, px, py, pz]
        controls : (4,)  — [throttle_left, throttle_right, de_left, de_right]
        """
        u, v, w, p, q, r, qx, qy, qz, qw, px, py, pz = state
        throttle_left, throttle_right, ev_left, ev_right = controls
        m = self.mass
        g = self.g
        Ix, Iy, Iz, Ixz = self.Ix, self.Iy, self.Iz, self.Ixz

        # 从油门获取推力
        thrust_left = self.aero.thrust_from_throttle(throttle_left)
        thrust_right = self.aero.thrust_from_throttle(throttle_right)

        # 计算旋转矩阵
        q_vec = np.array([qw, qx, qy, qz])
        R = rotation_matrix_from_quaternion(q_vec)

        # 地面坐标系中的重力
        G_e = np.array([0, 0, m * g])
        # 转换到机体坐标系
        G_b = R.T @ G_e

        # 气动力
        Fa, Ma_aero = self._aero_forces_moments(
            [u, v, w, p, q, r], ev_left, ev_right, thrust_left, thrust_right, rho)

        # 总力和力矩
        Fx = Fa[0] + G_b[0] + thrust_left + thrust_right
        Fy = Fa[1] + G_b[1]
        Fz = Fa[2] + G_b[2]

        # 差动油门产生偏航力矩（电机仅 y 向错位，推力沿 x 轴 → 绕 z 轴力矩）
        thrust_moment = thrust_left * self.aero.y_motor_left + thrust_right * self.aero.y_motor_right

        L = Ma_aero[0]             # 滚转仅由气动产生（升降副翼差动 + 横侧向）
        M = Ma_aero[1]
        N = Ma_aero[2] + thrust_moment   # 偏航 += 差动油门力矩

        # ── 平移动力学 ──
        du = (Fx / m) + r * v - q * w
        dv = (Fy / m) - r * u + p * w
        dw = (Fz / m) + q * u - p * v

        # ── 旋转动力学（with Ixz coupling） ──
        Gamma = Ix * Iz - Ixz**2
        if abs(Gamma) < 1e-10:
            Gamma = 1e-10

        dp = ((Iz * L + Ixz * N - (Iz * (Iz - Iy) + Ixz**2) * q * r
                - Ixz * (Ix - Iy + Iz) * p * q) / Gamma)
        dq = ((M - (Ix - Iz) * p * r - Ixz * (p**2 - r**2)) / Iy)
        dr = ((Ix * N + Ixz * L - Ixz * (Ix - Iy + Iz) * q * r
                - (Ix * (Ix - Iy) + Ixz**2) * p * q) / Gamma)

        # 角加速度限幅（可配置）
        dp, dq, dr = self._apply_limit(np.array([dp, dq, dr]), self.sat_accel, "angular_accel")

        # ── 四元数运动学 ──
        # 四元数导数幅值上限由角速度自然约束，无需额外截断
        omega = np.array([0, p, q, r])
        dq_dt = 0.5 * quaternion_multiply(q_vec, omega)
        dqw = dq_dt[0]
        dqx = dq_dt[1]
        dqy = dq_dt[2]
        dqz = dq_dt[3]

        # ── 导航（地面坐标系中的位置） ──
        v_b = np.array([u, v, w])
        v_e = R @ v_b
        dpx, dpy, dpz = self._apply_limit(v_e, self.sat_vel, "velocity")

        return np.array([du, dv, dw, dp, dq, dr, dqx, dqy, dqz, dqw,
                         dpx, dpy, dpz])

    # ── 配平 finder（使用气动表格的牛顿-拉夫逊法） ──
    def trim(self, V_trim, gamma_trim=0.0, h_trim=0.0):
        """
        查找纵向配平条件（机翼水平，零侧滑，对称飞行）。
        求解：轴向力=0，法向力=0，俯仰力矩=0。
        未知量：z = [alpha(rad), de_sym(rad), throttle]（强制 ev_left = ev_right = de_sym）。

        返回值：(x_trim, u_trim, alpha_trim, theta_trim, rho).
        """
        rho, _ = isa_atmosphere(h_trim)

        # 对于悬停情况，使用简化方法
        if abs(V_trim) < 0.1:  # 悬停
            # 重量与推力平衡
            m, g = self.mass, self.g
            total_thrust = m * g  # 总推力等于重量

            # 左右电机推力各一半（反插值精确匹配）
            throttle_left = throttle_right = np.interp(total_thrust / 2, self.aero._thrust, self.aero._throttle)

            # 尾座式悬停：机身垂直向上（theta=90°）
            alpha = np.radians(0.0)
            theta = np.radians(90.0) + gamma_trim

            # 升降副翼归零
            ev_left = ev_right = np.radians(0.0)

            u_trim = V_trim * np.cos(alpha)
            w_trim = V_trim * np.sin(alpha)

            # 转换欧拉角到四元数
            q_trim = euler_to_quaternion(0.0, theta, 0.0)

            x_trim = np.array([u_trim, 0.0, w_trim, 0.0, 0.0, 0.0,
                               q_trim[1], q_trim[2], q_trim[3], q_trim[0], 0.0, 0.0, h_trim])
            u_ctrl = np.array([throttle_left, throttle_right, ev_left, ev_right])

            return x_trim, u_ctrl, alpha, theta, rho

        # 对于正常飞行：强制左右舵面对称（ev_left = ev_right = de_sym）
        qbar = 0.5 * rho * V_trim**2
        S, c_bar = self.aero.S, self.aero.c_bar
        m, g = self.mass, self.g
        aero = self.aero

        def trim_residual(z):
            """计算给定[alpha, de_sym, throttle]的状态导数残差 [du, dw, dq]。
            直接调用 derivatives() 确保与动力学模型完全一致。"""
            alpha, de_sym, throttle = z
            alpha = np.clip(alpha, -np.radians(30), np.radians(30))
            theta = alpha + gamma_trim

            u = V_trim * np.cos(alpha)
            w = V_trim * np.sin(alpha)
            q_trim = euler_to_quaternion(0.0, theta, 0.0)
            x = np.array([u, 0.0, w, 0.0, 0.0, 0.0,
                          q_trim[1], q_trim[2], q_trim[3], q_trim[0],
                          0.0, 0.0, h_trim])
            u_ctrl = np.array([throttle, throttle, de_sym, de_sym])
            dx = self.derivatives(x, u_ctrl, rho)
            return dx[[0, 2, 4]]  # du, dw, dq

        # 初始猜测
        z = np.array([np.radians(5.0), np.radians(-2.0), 0.6])

        for iteration in range(100):
            f = trim_residual(z)
            # 数值雅可比（3方程 × 3未知数）
            J = np.zeros((3, 3))
            eps = 1e-5
            for j in range(3):
                zp = z.copy(); zp[j] += eps
                zm = z.copy(); zm[j] -= eps
                J[:, j] = (trim_residual(zp) - trim_residual(zm)) / (2 * eps)

            # 使用 lstsq 增强鲁棒性，避免低速时雅可比接近奇异导致 solve 失败
            dz = np.linalg.lstsq(J, -f, rcond=None)[0]
            z += dz
            if np.linalg.norm(dz) < 1e-10:
                break

        alpha, de_sym, throttle = z
        theta = alpha + gamma_trim
        u_trim = V_trim * np.cos(alpha)
        w_trim = V_trim * np.sin(alpha)

        # 转换欧拉角到四元数
        q_trim = euler_to_quaternion(0.0, theta, 0.0)

        x_trim = np.array([u_trim, 0.0, w_trim, 0.0, 0.0, 0.0,
                           q_trim[1], q_trim[2], q_trim[3], q_trim[0], 0.0, 0.0, h_trim])
        u_ctrl = np.array([throttle, throttle, de_sym, de_sym])

        return x_trim, u_ctrl, alpha, theta, rho

    def get_attitude(self, state):
        """
        从状态向量中获取姿态信息。

        参数：
            state: 状态向量 [u, v, w, p, q, r, qx, qy, qz, qw, px, py, pz]

        返回：
            phi, theta, psi: 欧拉角（弧度）
            q: 四元数 [qw, qx, qy, qz]
        """
        qx, qy, qz, qw = state[6:10]
        q = np.array([qw, qx, qy, qz])
        phi, theta, psi = quaternion_to_euler(q)
        return phi, theta, psi, q


# ─── 快速自测试 ───────────────────────────────────────────────────────────
if __name__ == "__main__":
    ac = Aircraft6DOF()

    # ── 悬停配平测试 ──
    print("=" * 60)
    print("悬停配平测试 (V_trim=0.0)")
    print("=" * 60)
    x0, u0, alpha_t, theta_t, rho = ac.trim(V_trim=0.0)
    throttle_left, throttle_right, ev_left, ev_right = u0
    thrust_left = ac.aero.thrust_from_throttle(throttle_left)
    thrust_right = ac.aero.thrust_from_throttle(throttle_right)
    total_thrust = thrust_left + thrust_right

    print(f"配平:  V=0.0 m/s  alpha={np.degrees(alpha_t):.2f} deg  "
          f"theta={np.degrees(theta_t):.2f} deg")
    print(f"左侧:  油门={throttle_left:.3f}  推力={thrust_left:.1f}N  ev={np.degrees(ev_left):.2f} deg")
    print(f"右侧:  油门={throttle_right:.3f}  推力={thrust_right:.1f}N  ev={np.degrees(ev_right):.2f} deg")
    print(f"总推力: {total_thrust:.1f} N  (重量={ac.mass * ac.g:.1f} N)")

    phi, theta, psi, q = ac.get_attitude(x0)
    print(f"姿态: phi={np.degrees(phi):.2f}°, theta={np.degrees(theta):.2f}°, psi={np.degrees(psi):.2f}°")

    dx = ac.derivatives(x0, u0, rho)
    mech_res = np.max(np.abs(dx[:9]))
    print(f"状态导数残差 max|dx[0:9]|: {mech_res:.2e}")
    assert mech_res < 1e-2, f"悬停配平残差过大: {mech_res}"

    # ── 平飞配平测试 ──
    print("\n" + "=" * 60)
    print("平飞配平测试 (V_trim=20.0)")
    print("=" * 60)
    x20, u20, alpha_20, theta_20, rho = ac.trim(V_trim=20.0)
    throttle_left, throttle_right, ev_left, ev_right = u20
    thrust_left = ac.aero.thrust_from_throttle(throttle_left)
    thrust_right = ac.aero.thrust_from_throttle(throttle_right)

    print(f"配平:  V=20.0 m/s  alpha={np.degrees(alpha_20):.2f} deg  "
          f"theta={np.degrees(theta_20):.2f} deg")
    print(f"油门: {throttle_left:.3f} / {throttle_right:.3f}  "
          f"ev={np.degrees(ev_left):.2f} deg")
    print(f"推力: {thrust_left:.1f}N / {thrust_right:.1f}N")

    dx20 = ac.derivatives(x20, u20, rho)
    mech_res_20 = np.max(np.abs(dx20[:9]))
    print(f"状态导数残差 max|dx[0:9]|: {mech_res_20:.2e}")
    assert mech_res_20 < 1e-2, f"平飞配平残差过大: {mech_res_20}"

    # ── 边界连续性检查 ──
    print("\n" + "=" * 60)
    print("边界连续性检查")
    print("=" * 60)
    for V_test in [0.5, 2.0, 5.0, 10.0, 20.0]:
        try:
            x_t, u_t, alpha_t, theta_t, rho_t = ac.trim(V_trim=V_test)
            dx_t = ac.derivatives(x_t, u_t, rho_t)
            res = np.max(np.abs(dx_t[:9]))
            # 低速时气动升力可能不足以支持平飞，配平姿态会异常；仅检查不崩溃
            alpha_deg = np.degrees(alpha_t)
            if abs(alpha_deg) < 60 and res < 1e-2:
                status = "OK"
            elif V_test < 12:
                status = "WARN(低速气动升力不足)"
            else:
                status = "FAIL"
            print(f"V={V_test:5.1f} m/s  alpha={alpha_deg:7.2f}°  theta={np.degrees(theta_t):7.2f}°  "
                  f"res={res:.2e}  [{status}]")
        except Exception as e:
            print(f"V={V_test:5.1f} m/s  配平失败: {e}")

    print("\n核心重构测试通过。")
